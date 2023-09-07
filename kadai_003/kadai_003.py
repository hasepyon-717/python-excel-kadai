import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill 

# 処理対象のファイル名のリストを格納
file_lists = ['2022_年間売上表.xlsx', '2023_年間売上表.xlsx']

# 結合後のデータを格納するための空のデータフレームを作成
df = pd.DataFrame()

file_data = [
    pd.read_excel(file_list, sheet_name = 'Sheet1')
    for file_list in file_lists

]

all_data = pd.concat(file_data, ignore_index=True)


    
grouped = all_data.groupby(['商品', '売上年'])['金額（千円）'].sum().reset_index()

# Excelファイル名
outputfile='売上集計表.xlsx'
 
# Excelファイルに書き込む
grouped.to_excel(outputfile,  index=False)

book = load_workbook(outputfile)

sheet = book.active
for cell in sheet[1]:
    cell.fill = PatternFill(patternType = 'solid', fgColor='F2F2F2')

book.save(outputfile)

